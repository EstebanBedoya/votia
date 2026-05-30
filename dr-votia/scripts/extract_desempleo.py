"""Extract key labour-market figures from the DANE GEIH workbook into curated
markdown.

The raw workbook has 21 sheets with hundreds of monthly columns (~970K cells) —
unsuitable for cell-level embedding. This distils the headline indicators of the
national and 13-cities sheets into annual averages + the latest value per
indicator: accurate (computed from the real cells), bounded (~dozens of blocks),
and self-describing, so each becomes a clean retrievable chunk via the markdown
ingestion path.

Run:  uv run python scripts/extract_desempleo.py
"""

from __future__ import annotations

from pathlib import Path
from statistics import mean

from openpyxl import load_workbook

DATA = Path(__file__).resolve().parents[2] / "dr-contexto-data" / "datasets_nacionales"
XLSX = DATA / "dane_desempleo_geih_abr2026.xlsx"
OUT = DATA / "dane_desempleo_resumen.md"

# Sheets that hold headline indicators as rows × monthly columns.
TARGET_SHEETS = {
    "Total nacional": "Total nacional",
    "Total 13 ciudades A.M.": "13 ciudades y áreas metropolitanas",
}
_MONTHS = ("ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic")


def _is_year(value: object) -> bool:
    text = str(value).strip() if value is not None else ""
    return text[:4].isdigit() and len(text[:4]) == 4 and text[:4].startswith(("19", "20"))


def _find_month_row(rows: list[tuple[object, ...]]) -> int:
    for idx, row in enumerate(rows):
        cells = {str(c).strip().lower()[:3] for c in row if c is not None}
        if "ene" in cells and "feb" in cells:
            return idx
    raise ValueError("No month header row found")


def _column_periods(year_row: tuple[object, ...], month_row: tuple[object, ...]) -> dict[int, str]:
    """Map each column index to its 'YYYY' year, forward-filling sparse years."""
    periods: dict[int, str] = {}
    current_year = ""
    width = max(len(year_row), len(month_row))
    for col in range(1, width):
        year_cell = year_row[col] if col < len(year_row) else None
        if _is_year(year_cell):
            current_year = str(year_cell).strip()[:4]
        month_cell = month_row[col] if col < len(month_row) else None
        is_month = month_cell is not None and str(month_cell).strip().lower()[:3] in _MONTHS
        if current_year and is_month:
            periods[col] = current_year
    return periods


def _annual_averages(values: tuple[object, ...], col_year: dict[int, str]) -> dict[str, float]:
    buckets: dict[str, list[float]] = {}
    for col, year in col_year.items():
        if col < len(values) and isinstance(values[col], int | float):
            buckets.setdefault(year, []).append(float(values[col]))
    return {year: round(mean(vals), 1) for year, vals in buckets.items() if vals}


def _latest(
    values: tuple[object, ...], col_year: dict[int, str], month_row: tuple[object, ...]
) -> str:
    for col in sorted(col_year, reverse=True):
        if col < len(values) and isinstance(values[col], int | float):
            month = str(month_row[col]).strip() if col < len(month_row) else ""
            return f"{month} {col_year[col]}: {round(float(values[col]), 1)}"
    return "s/d"


def _unit(annual: dict[str, float]) -> str:
    """Rates sit well below 100; population counts are in thousands."""
    values = sorted(annual.values())
    median = values[len(values) // 2]
    return "miles de personas" if median > 100 else "%"


def _sheet_markdown(label: str, rows: list[tuple[object, ...]]) -> list[str]:
    month_idx = _find_month_row(rows)
    col_year = _column_periods(rows[month_idx - 1], rows[month_idx])
    blocks: list[str] = []
    seen: set[str] = set()  # the sheet stacks the same indicators more than once
    for row in rows[month_idx + 1 :]:
        name = str(row[0]).strip() if row and row[0] is not None else ""
        if len(name) < 4 or name in seen:
            continue
        annual = _annual_averages(row, col_year)
        if not annual:
            continue
        seen.add(name)
        unit = _unit(annual)
        series = " | ".join(f"{y}: {v}" for y, v in sorted(annual.items()))
        latest = _latest(row, col_year, rows[month_idx])
        blocks.append(
            f"### {label} — {name} ({unit})\n"
            f"Promedios anuales (DANE GEIH): {series}\n"
            f"Último dato disponible — {latest} {unit}\n"
        )
    return blocks


def main() -> None:
    wb = load_workbook(str(XLSX), read_only=True, data_only=True)
    out: list[str] = [
        "# Desempleo y mercado laboral — Colombia (DANE, GEIH)",
        "",
        "Cifras clave extraídas de la Gran Encuesta Integrada de Hogares (serie "
        "mensual 2001–2026). Promedios anuales calculados sobre los datos mensuales. "
        "Indicadores: Tasa Global de Participación (TGP), Tasa de Ocupación (TO), "
        "Tasa de Desocupación (TD, desempleo), Tasa de Subocupación (TS).",
        "",
    ]
    for sheet, label in TARGET_SHEETS.items():
        rows = list(wb[sheet].iter_rows(values_only=True))
        out.append(f"## {label}\n")
        out.extend(_sheet_markdown(label, rows))
    wb.close()

    OUT.write_text("\n".join(out), encoding="utf-8")
    print(f"Escrito: {OUT}")
    print(f"Bloques de indicadores: {sum(1 for line in out if line.startswith('### '))}")
    print(f"Tamaño: {OUT.stat().st_size} bytes")


if __name__ == "__main__":
    main()
