"""XLSX reader adapter (openpyxl). Implements the DocumentReader port.

Statistical spreadsheets (DANE, BanRep) are often very WIDE — a single sheet can
carry hundreds of columns (a monthly time series per indicator). Dumping a row as
one flat line and letting the chunker slice it blindly produces thousands of
headerless, unsearchable fragments.

Instead, this reader extracts ONE OR MORE self-describing fragments per data row:
each carries ``Hoja · preámbulo · etiqueta-de-fila`` as a prefix, followed by a
bounded block of ``columna: valor`` pairs. When a row is too wide, it is split
into several fragments and the prefix is repeated, so every chunk stays
retrievable and meaningful on its own. No data is dropped.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from dr_votia.domain.models import Fragment

Row = tuple[object, ...]
MAX_HEADER_SCAN = 25  # banner rows always sit near the top
MAX_FRAGMENT_CHARS = 700  # keep each fragment under the chunk budget (≈800)


class XlsxReader:
    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in {".xlsx", ".xlsm"}

    def read(self, path: Path) -> list[Fragment]:
        workbook = load_workbook(str(path), read_only=True, data_only=True)
        fragments: list[Fragment] = []
        for sheet in workbook.worksheets:
            rows: list[Row] = [
                r for r in sheet.iter_rows(values_only=True) if any(c is not None for c in r)
            ]
            fragments.extend(_sheet_fragments(str(sheet.title), rows))
        workbook.close()
        return fragments


def _sheet_fragments(title: str, rows: list[Row]) -> list[Fragment]:
    if not rows:
        return []

    scan = rows[:MAX_HEADER_SCAN]
    fill_counts = [sum(1 for c in r if c is not None) for r in scan]
    header_idx = max(range(len(scan)), key=lambda i: fill_counts[i])

    # No tabular structure (e.g. an index or methodology sheet): pack plain lines.
    if fill_counts[header_idx] < 2:
        lines = [_join_row(r) for r in rows]
        return _pack(title, lines)

    headers = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    preamble = " · ".join(str(c).strip() for r in rows[:header_idx] for c in r if c is not None)
    context = f"{title} · {preamble}".strip(" ·") if preamble else title

    fragments: list[Fragment] = []
    for row in rows[header_idx + 1 :]:
        fragments.extend(_row_fragments(context, row, headers))
    return fragments


def _row_fragments(context: str, row: Row, headers: list[str]) -> list[Fragment]:
    label = str(row[0]).strip() if row and row[0] is not None else ""
    prefix = f"{context} · {label}".strip(" ·") if label else context

    pairs: list[str] = []
    for i in range(1, len(row)):
        value = row[i]
        if value is None:
            continue
        key = headers[i] if i < len(headers) and headers[i] else ""
        pairs.append(f"{key}: {value}" if key else str(value))

    return _pack(prefix, pairs, joiner=" | ", sep=" — ")


def _pack(prefix: str, items: list[str], *, joiner: str = "\n", sep: str = "\n") -> list[Fragment]:
    """Group items into fragments, each prefixed with ``prefix`` and ≤ budget."""
    if not items:
        return [Fragment(text=prefix)] if len(prefix) > 20 else []

    fragments: list[Fragment] = []
    base = len(prefix) + len(sep)
    current: list[str] = []
    length = base
    for item in items:
        if current and length + len(item) + len(joiner) > MAX_FRAGMENT_CHARS:
            fragments.append(Fragment(text=f"{prefix}{sep}{joiner.join(current)}"))
            current, length = [], base
        current.append(item)
        length += len(item) + len(joiner)
    if current:
        fragments.append(Fragment(text=f"{prefix}{sep}{joiner.join(current)}"))
    return fragments


def _join_row(row: Row) -> str:
    return " | ".join(str(c).strip() for c in row if c is not None)
